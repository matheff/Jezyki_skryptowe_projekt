
class ExcelExporter:
    CURRENCY_FORMAT = "#,##0.00"

    def export(self, dataset, statistics, products, filepath: str) -> None:
        try:
            import openpyxl
        except ImportError:
            raise ImportError( "Biblioteka openpyxl nie jest zainstalowana. "
                                "Uruchom: pip install openpyxl")
        
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            self.create_transaction_sheet(wb, dataset)
            self.create_statistics_sheet(wb, statistics)
            self.create_product_sheet(wb, products)

            wb.save(filepath)
        except OSError as e:
            raise IOError(f"Nie można zapisać pliku Excel: {filepath}") from e

    def create_transaction_sheet(self, wb, dataset):
    
        sheetTransaction = wb.create_sheet("Transakcje")

        headers = ["Data", "Produkt", "Kategoria", "Ilość", "Cena jedn.", "Wartość", "Sprzedawca", "Region"]
        
        sheetTransaction.append(headers)

        self.style_header(sheetTransaction)
        sheetTransaction.freeze_panes="A2"

        records = sorted(dataset, key=lambda r: r.date)

        for r in records:
            sheetTransaction.append([
                r.date,
                r.product.name,
                r.product.category,
                r.quantity,
                r.product.unit_price,
                r.total_price(),
                r.seller,
                r.region_code
            ])

        self.set_currency_format(sheetTransaction, ["E", "F"], self.CURRENCY_FORMAT)

        self.auto_size(sheetTransaction)

    def create_statistics_sheet(self, wb, statistics):
        
        sheetStatistics = wb.create_sheet("Statystyki")

        row = 1

        row = self.write_section_title(sheetStatistics, row, 1, "PODSUMOWANIE")

        row = self.write_section_cell(sheetStatistics, row, 1, "Łączny przychód", statistics.total_revenue(), currency=True)
        row = self.write_section_cell(sheetStatistics, row, 1, "Liczba transakcji", statistics.transaction_count())
        row = self.write_section_cell(sheetStatistics, row, 1, "Średnia wartość", statistics.average_revenue() if statistics.transaction_count() > 0 else 0, currency=True)

        row += 2

        row = self.write_section_title(sheetStatistics, row, 2, "PRZYCHÓD WG. KATEGORII")
        row = self.write_section_cell_dictionary(sheetStatistics, row, 1, statistics.revenue_by_category(), currency=True)

        row += 2

        row = self.write_section_title(sheetStatistics, row, 2, "PRZYCHÓD WG. SPRZEDAWCY")
        row = self.write_section_cell_dictionary(sheetStatistics, row, 1, statistics.revenue_by_seller(), currency=True)

        row += 2

        row = self.write_section_title(sheetStatistics, row, 2, "TOP 5 PRODUKTÓW")
        row = self.write_section_cell_dictionary(sheetStatistics, row, 1, statistics.top_revenue_by_product(), currency=True)

        row += 2

        row = self.write_section_title(sheetStatistics, row, 2, "PODSUMOWANIE MIESIĘCZNE")
        row = self.write_section_cell_dictionary(sheetStatistics, row, 1, statistics.revenue_by_date(), currency=True)

        self.auto_size(sheetStatistics)

    def create_product_sheet(self, wb, products_dict):
        sheetProduct = wb.create_sheet("Produkty")
        
        headers = ["ID", "Nazwa", "Kategoria", "Cena (PLN)"]

        sheetProduct.append(headers)

        self.style_header(sheetProduct)

        sorted_products = sorted(products_dict.values(), key=lambda r: r.product_id)

        for p in sorted_products:
            sheetProduct.append([
                p.product_id,
                p.name,
                p.category,
                p.unit_price
            ])

        self.set_currency_format(sheetProduct, ["D"], self.CURRENCY_FORMAT)
        self.auto_size(sheetProduct)


    def auto_size(self, sheet):
        for column_cells in sheet.columns:
            length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    def write_section_title(self, sheet, row, column, title):
        cell = sheet.cell(row=row, column=column, value=title) #to się i tak wykona bo w pythonie jest tak że jak jest równa się to wykonuje się to po prawej stronie, pobiera wynik i dopiero przypisuje się do zmiennej
        self.make_bold(cell, size=12)
        return row + 1

    def write_section_cell(self, sheet, row, column, label, value, currency=False):
        sheet.cell(row=row, column=column, value=label)
        cell_value = sheet.cell(row=row, column=column + 1, value=value)

        if currency:
            cell_value.number_format = self.CURRENCY_FORMAT

        return row + 1
    
    def write_section_cell_dictionary(self, sheet, row, column, data: dict, currency=False):
        for key, value in data.items():
            sheet.cell(row=row, column=column, value=key)
            cell_value = sheet.cell(row=row, column=column + 1, value=value)

            if currency:
                cell_value.number_format = self.CURRENCY_FORMAT

            row += 1
        return row
    
    def set_currency_format(self, sheet, columns, format):
        for col in columns:
            for cell in sheet[col][1:]: # [1:] Pominięcie wiersza nagłówkowego
                cell.number_format = format

    def style_header(self, sheet):
        for cell in sheet[1]:
            self.make_bold(cell)

    def make_bold(self, cell, size=None): 
        from openpyxl.styles import Font 

        cell.font = Font(bold=True, size=size)